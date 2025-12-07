import io
import re
import glob
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st


st.set_page_config(
    page_title="Screener Aggregator", 
    layout="wide",
    initial_sidebar_state="expanded"
)


# ---------- Normalization helpers ----------
def normalize_label(text: str) -> str:
    """Lowercase alphanumeric key for matching columns/criteria."""
    if text is None or (isinstance(text, float) and np.isnan(text)):
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(text).lower())


def detect_header_row(df: pd.DataFrame) -> int:
    """Return first row index containing 'ticker'; default to 0."""
    mask = df.apply(
        lambda row: row.astype(str).str.contains("ticker", case=False, na=False)
    ).any(axis=1)
    idx = list(df.index[mask])
    return idx[0] if idx else 0


def to_bytes_buffer(file_like) -> io.BytesIO:
    """Ensure we can rewind uploads/paths for multiple reads."""
    if isinstance(file_like, (str, Path)):
        path = Path(file_like)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_like}")
        data = path.read_bytes()
    else:
        data = file_like.read()
    return io.BytesIO(data)


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize column names and keep them unique."""
    seen = set()
    new_cols = []
    for col in df.columns:
        key = normalize_label(col) or "col"
        base = key
        counter = 1
        while key in seen:
            counter += 1
            key = f"{base}_{counter}"
        seen.add(key)
        new_cols.append(key)
    df = df.copy()
    df.columns = new_cols
    return df


def resolve_ticker_column(df: pd.DataFrame) -> Optional[str]:
    """Pick the best ticker-like column."""
    for cand in ("ticker", "fullticker", "symbol"):
        if cand in df.columns:
            return cand
    fallback = [c for c in df.columns if "ticker" in c]
    return fallback[0] if fallback else None


# ---------- Criteria parsing ----------
@st.cache_data(show_spinner=False)
def load_ratio_criteria(file_like) -> pd.DataFrame:
    """Read criteria ranges; expects the table that starts with 'Factor'."""
    try:
        buffer = to_bytes_buffer(file_like)
        raw = pd.read_excel(buffer, header=None)
        start_rows = raw[raw.eq("Factor").any(axis=1)].index
        if len(start_rows) == 0:
            return pd.DataFrame()
        start = start_rows[0]
        block = raw.iloc[start:, 2:7]
        block.columns = ["factor", "description", "good", "avg", "bad"]
        block = block.dropna(subset=["description"])
        block["norm_desc"] = block["description"].apply(normalize_label)
        return block.reset_index(drop=True)
    except (FileNotFoundError, Exception) as e:
        # Return empty DataFrame if file can't be loaded
        return pd.DataFrame()


# ---------- Screener ingestion ----------
@st.cache_data(show_spinner=False)
def load_screener(file_like) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """Read one screener (Excel/CSV), detect header row, clean, normalize."""
    name = getattr(file_like, "name", str(file_like))
    ext = Path(name).suffix.lower()
    excel_exts = {".xlsx", ".xls"}
    buffer = to_bytes_buffer(file_like)

    try:
        if ext in excel_exts:
            raw = pd.read_excel(buffer, header=None)
        else:
            raw = pd.read_csv(buffer, header=None)
    except Exception as exc:  # pragma: no cover - defensive
        return None, f"{name}: could not read file ({exc})"

    header_idx = detect_header_row(raw)
    if hasattr(buffer, "seek"):
        buffer.seek(0)

    try:
        if ext in excel_exts:
            df = pd.read_excel(buffer, header=header_idx)
        else:
            df = pd.read_csv(buffer, header=header_idx)
    except Exception as exc:  # pragma: no cover - defensive
        return None, f"{name}: could not parse header ({exc})"

    df = df.dropna(how="all")
    df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed")]
    df = standardize_columns(df)

    ticker_col = resolve_ticker_column(df)
    if not ticker_col:
        return None, f"{name}: no ticker column found"

    df = df.dropna(subset=[ticker_col])
    df[ticker_col] = df[ticker_col].astype(str).str.strip()
    df = df[df[ticker_col] != ""]
    df = df.rename(columns={ticker_col: "ticker"})
    return df.reset_index(drop=True), None


def merge_screeners(dfs: List[pd.DataFrame]) -> pd.DataFrame:
    combined = None
    for df in dfs:
        if combined is None:
            combined = df
        else:
            combined = combined.merge(df, on="ticker", how="outer", suffixes=("", "_dup"))
    if combined is None:
        return pd.DataFrame()
    dup_cols = [c for c in combined.columns if c.endswith("_dup")]
    combined = combined.drop(columns=dup_cols)
    return combined


def safe_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """Convert any problematic cells to strings to keep Arrow happy.
    
    Handles:
    - Binary/bytes-like cells
    - Mixed-type columns (e.g., numeric columns with 'NM', 'N/A' strings)
    """
    cleaned = df.copy()
    for col in cleaned.columns:
        if cleaned[col].dtype == object:
            # Check if column has mixed types (numbers + strings like 'NM')
            has_numeric = cleaned[col].apply(
                lambda v: isinstance(v, (int, float, np.integer, np.floating)) and not pd.isna(v)
            ).any()
            has_non_numeric_str = cleaned[col].apply(
                lambda v: isinstance(v, str) and v.strip() not in ('', 'nan', 'NaN') 
                          and not v.replace('.', '', 1).replace('-', '', 1).replace('%', '').replace(',', '').isdigit()
            ).any()
            
            if has_numeric and has_non_numeric_str:
                # Convert entire column to string to avoid Arrow serialization issues
                cleaned[col] = cleaned[col].apply(
                    lambda v: '' if pd.isna(v) else str(v)
                )
            else:
                # Original handling for bytes
                cleaned[col] = cleaned[col].apply(
                    lambda v: v.decode() if isinstance(v, (bytes, bytearray)) else v
                )
    return cleaned


# ---------- Scoring helpers ----------
def coerce_numeric(value) -> Optional[float]:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = str(value).replace(",", "").replace("%", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def parse_range(cond: str) -> Tuple[Optional[float], Optional[float], Optional[str]]:
    """Return (min, max, operator) where operator in {'<', '>', '<=', '>='} or None."""
    if not cond or cond.lower() in {"x", "nan"}:
        return None, None, None
    c = cond.strip().lower()
    c = c.replace("%", "").replace(" ", "")

    # ranges like 5-15 or 1.5-3
    if "-" in c and not c.startswith("-"):
        parts = c.split("-")
        if len(parts) == 2:
            lo, hi = parts
            try:
                return float(lo), float(hi), "between"
            except ValueError:
                return None, None, None

    # inequality
    for op in (">=", "<=", ">", "<"):
        if c.startswith(op):
            try:
                return float(c[len(op) :]), None, op
            except ValueError:
                return None, None, None

    # >= as textual (>=4 etc already handled)
    if c.startswith(">="):
        try:
            return float(c[2:]), None, ">="
        except ValueError:
            return None, None, None

    if c in {"positive", "pos"}:
        return 0.0, None, ">"

    return None, None, None


def check_condition(value, cond: str) -> Optional[bool]:
    """Return True/False if evaluable, else None for non-numeric/unsupported."""
    if cond is None or (isinstance(cond, float) and np.isnan(cond)):
        return None
    cond_text = str(cond)
    value_num = coerce_numeric(value)

    # Handle relational text we cannot evaluate
    tokens = cond_text.lower()
    relational_keywords = ("3y", "5y", "tbills", "bond", "cfo", "ccfo", "net profit")
    if any(tok in tokens for tok in relational_keywords):
        return None

    if tokens in {"x", "nan", "", "n/a"}:
        return None

    lo, hi, op = parse_range(cond_text)
    if value_num is None or op is None:
        return None

    if op == "between" and lo is not None and hi is not None:
        return lo <= value_num <= hi
    if op == ">":
        return value_num > lo
    if op == "<":
        return value_num < lo
    if op == ">=":
        return value_num >= lo
    if op == "<=":
        return value_num <= lo
    return None


def classify_value(value, good_cond, avg_cond, bad_cond) -> str:
    if pd.isna(value):
        return "Missing"

    if (res := check_condition(value, good_cond)) is True:
        return "Good"
    if (res := check_condition(value, avg_cond)) is True:
        return "Average"
    if (res := check_condition(value, bad_cond)) is True:
        return "Bad"

    # If none of the conditions were evaluable or matched
    if all(check_condition(value, c) is None for c in (good_cond, avg_cond, bad_cond)):
        return "No rule"
    return "Unknown"


def score_dataframe(df: pd.DataFrame, criteria: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    if df.empty or criteria.empty:
        return df, []

    columns_lookup = {normalize_label(c): c for c in df.columns if c != "ticker"}
    scored = df.copy()
    score_columns: List[str] = []
    factor_columns: Dict[str, List[str]] = {}

    for _, row in criteria.iterrows():
        norm = row.get("norm_desc", "")
        target_col = columns_lookup.get(norm)
        if not target_col:
            continue
        label = str(row["description"])
        score_col = f"{label} (score)"
        scored[score_col] = scored[target_col].apply(
            lambda v: classify_value(v, row.get("good"), row.get("avg"), row.get("bad"))
        )
        score_columns.append(score_col)
        factor = str(row.get("factor") or "Other")
        factor_columns.setdefault(factor, []).append(score_col)

    score_to_int = {"Good": 2, "Average": 1, "Bad": 0}
    total_cols = []
    for factor, cols in factor_columns.items():
        total_col = f"{factor.lower()} scores total"
        scored[total_col] = scored[cols].apply(
            lambda row: row.map(score_to_int).sum(), axis=1
        )
        total_cols.append(total_col)

    if total_cols:
        scored["Total Financial Scores"] = scored[total_cols].sum(axis=1)
        scored["Total Ranking"] = (
            scored["Total Financial Scores"].rank(ascending=False, method="dense").astype(int)
        )
        scored = scored.sort_values("Total Financial Scores", ascending=False)

    return scored.reset_index(drop=True), score_columns


# ---------- Download helpers ----------
def to_excel_bytes(
    df: pd.DataFrame, 
    score_columns: Optional[List[str]] = None,
    raw_df: Optional[pd.DataFrame] = None,
    criteria_df: Optional[pd.DataFrame] = None
) -> bytes:
    """Export DataFrame to professional Excel with multiple sheets:
    - Summary: Pivot-style aggregation by factor
    - Scored Results: Full data with color coding
    - Raw Data: Original merged data
    """
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output = io.BytesIO()
    
    # Color definitions
    score_colors = {
        "Good": "C8E6C9",      # Light green
        "Average": "FFF9C4",   # Light yellow
        "Bad": "FFCDD2",       # Light red
        "Missing": "E0E0E0",   # Light gray
        "No rule": "E0E0E0",   # Light gray
        "Unknown": "E0E0E0",   # Light gray
    }
    
    header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    def style_header_row(worksheet, num_cols):
        """Apply professional styling to header row."""
        for col_idx in range(1, num_cols + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        worksheet.freeze_panes = 'A2'
    
    def auto_fit_columns(worksheet, df_cols, min_width=10, max_width=50):
        """Auto-fit column widths based on content."""
        for idx, col_name in enumerate(df_cols, start=1):
            col_letter = get_column_letter(idx)
            # Calculate width based on header and sample values
            header_len = len(str(col_name))
            worksheet.column_dimensions[col_letter].width = min(max(header_len + 2, min_width), max_width)
    
    def apply_borders(worksheet, num_rows, num_cols):
        """Apply thin borders to all data cells."""
        for row_idx in range(2, num_rows + 2):
            for col_idx in range(1, num_cols + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ---------- Sheet 1: Summary (Pivot-style) ----------
        if score_columns and criteria_df is not None and not criteria_df.empty:
            # Create summary with counts of Good/Avg/Bad per factor
            summary_data = []
            
            # Get unique factors from criteria
            factors = criteria_df['factor'].dropna().unique().tolist()
            
            for _, row in df.iterrows():
                ticker = row.get('ticker', 'Unknown')
                row_data = {'Ticker': ticker}
                
                total_good = 0
                total_avg = 0
                total_bad = 0
                
                for factor in factors:
                    factor_cols = [c for c in score_columns if c in df.columns]
                    factor_criteria = criteria_df[criteria_df['factor'] == factor]['description'].tolist()
                    
                    good_count = 0
                    avg_count = 0
                    bad_count = 0
                    
                    for col in factor_cols:
                        # Check if this score column corresponds to this factor
                        col_base = col.replace(' (score)', '')
                        if col_base in factor_criteria:
                            val = row.get(col, '')
                            if val == 'Good':
                                good_count += 1
                            elif val == 'Average':
                                avg_count += 1
                            elif val == 'Bad':
                                bad_count += 1
                    
                    row_data[f'{factor} Good'] = good_count
                    row_data[f'{factor} Avg'] = avg_count
                    row_data[f'{factor} Bad'] = bad_count
                    total_good += good_count
                    total_avg += avg_count
                    total_bad += bad_count
                
                row_data['Total Good'] = total_good
                row_data['Total Avg'] = total_avg
                row_data['Total Bad'] = bad_count
                row_data['Score'] = row.get('Total Financial Scores', 0)
                row_data['Rank'] = row.get('Total Ranking', '')
                
                summary_data.append(row_data)
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
            
            ws_summary = writer.sheets["Summary"]
            style_header_row(ws_summary, len(summary_df.columns))
            auto_fit_columns(ws_summary, summary_df.columns)
            apply_borders(ws_summary, len(summary_df), len(summary_df.columns))
        
        # ---------- Sheet 2: Scored Results ----------
        df.to_excel(writer, index=False, sheet_name="Scored Results")
        ws_scored = writer.sheets["Scored Results"]
        
        style_header_row(ws_scored, len(df.columns))
        auto_fit_columns(ws_scored, df.columns)
        
        # Apply color coding to score columns
        if score_columns:
            df_columns = list(df.columns)
            for col_name in score_columns:
                if col_name not in df_columns:
                    continue
                col_idx = df_columns.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                
                for row_idx, value in enumerate(df[col_name], start=2):
                    cell = ws_scored[f"{col_letter}{row_idx}"]
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    if pd.isna(value):
                        fill_color = score_colors.get("Missing", "FFFFFF")
                    else:
                        val_str = str(value).strip()
                        fill_color = score_colors.get(val_str, "FFFFFF")
                    
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # ---------- Sheet 3: Raw Data ----------
        if raw_df is not None and not raw_df.empty:
            raw_df.to_excel(writer, index=False, sheet_name="Raw Data")
            ws_raw = writer.sheets["Raw Data"]
            style_header_row(ws_raw, len(raw_df.columns))
            auto_fit_columns(ws_raw, raw_df.columns)
    
    return output.getvalue()


# ---------- UI ----------
# Add dark mode CSS support and clean up UI
st.markdown("""
<style>
    /* Hide status messages and spinners on main page */
    [data-testid="stStatus"],
    [data-testid="stSpinner"] {
        display: none !important;
    }
    
    /* Dark mode support - works with Streamlit's theme system */
    .stApp {
        background-color: var(--background-color, #ffffff);
    }
    
    /* Ensure proper contrast in dark mode */
    [data-testid="stDataFrame"] {
        background-color: transparent;
    }
    
    /* Style improvements */
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    
    /* Better spacing for main content */
    h1 {
        margin-bottom: 1rem;
    }
    
    /* Hide any processing indicators */
    .stSpinner > div {
        display: none !important;
    }
</style>
""", unsafe_allow_html=True)

st.title("📊 Screener Aggregator")

with st.sidebar:
    st.header("⚙️ Configuration")
    
    # Criteria file upload
    criteria_upload = st.file_uploader(
        "Upload Ratio Criteria File (Optional)",
        type=["xlsx", "xls"],
        help="Upload the Ratio_Criteria.xlsx file. If not provided, the app will try to find it in sample_files/."
    )
    
    screener_uploads = st.file_uploader(
        "Upload Screener Files", 
        type=["xlsx", "xls", "csv"], 
        accept_multiple_files=True,
        help="Upload one or more screener files (Excel/CSV) to analyze"
    )

    st.markdown("---")
    st.caption(
        "💡 **Note:** We automatically detect the header row containing 'Ticker' and normalize all columns for merging."
    )


# Load criteria file - try multiple locations
selected_files = list(screener_uploads) if screener_uploads else []

criteria_df = pd.DataFrame()
criteria_source = None

# First, try uploaded file
if criteria_upload:
    criteria_df = load_ratio_criteria(criteria_upload)
    criteria_source = criteria_upload

# If no upload or empty, try local paths
if criteria_df.empty:
    possible_paths = [
        Path("sample_files/Ratio_Criteria.xlsx"),
        Path("Ratio_Criteria.xlsx"),
        Path("./sample_files/Ratio_Criteria.xlsx"),
    ]
    
    for path in possible_paths:
        try:
            if path.exists():
                criteria_df = load_ratio_criteria(path)
                if not criteria_df.empty:
                    criteria_source = path
                    break
        except Exception:
            continue

# If still empty, show error
if criteria_df.empty:
    st.error(
        "❌ **Could not locate criteria table.**\n\n"
        "Please upload the Ratio_Criteria.xlsx file in the sidebar, or ensure it exists in the sample_files/ directory."
    )
    st.stop()

if not selected_files:
    st.info("👆 Upload one or more screener files (Excel/CSV) in the sidebar to begin analysis.")
    st.stop()


# Load screeners (processing happens silently)
loaded_screeners: List[pd.DataFrame] = []
load_errors: List[str] = []

for f in selected_files:
    df, err = load_screener(f)
    if err:
        load_errors.append(err)
    elif df is not None:
        loaded_screeners.append(df)

if load_errors:
    st.warning("⚠️ Some files were skipped:\n- " + "\n- ".join(load_errors))

if not loaded_screeners:
    st.error("❌ No valid screener files to process.")
    st.stop()

# Process data silently
combined = merge_screeners(loaded_screeners)
scored, score_cols = score_dataframe(combined, criteria_df)
    
if scored.empty:
    st.error("❌ Could not score data. Check that tickers and numeric columns exist.")
    st.stop()

# Main page - Only show final results
st.header("🎯 Final Results")
st.caption("Ranked and scored stocks with color-coded ratings (Good/Average/Bad)")

color_map = {
    "Good": "#2e7d32",
    "Average": "#f9a825",
    "Bad": "#c62828",
    "Missing": "#424242",
    "No rule": "#424242",
    "Unknown": "#424242",
}


def highlight_scores(val):
    """Apply color coding to score values. Handles NaN and ensures scalar input."""
    try:
        # Ensure we're working with a scalar, not a Series
        if isinstance(val, (pd.Series, pd.DataFrame)):
            return ""
        # Handle NaN/None
        if val is None or (isinstance(val, float) and (pd.isna(val) or np.isnan(val))):
            return ""
        # Convert to string and get color
        val_str = str(val).strip()
        if not val_str or val_str.lower() == "nan":
            return ""
        color = color_map.get(val_str, "")
        return f"background-color: {color}" if color else ""
    except Exception:
        # If anything goes wrong, return empty string (no styling)
        return ""


score_columns_present = [c for c in score_cols if c in scored.columns]
if score_columns_present:
    try:
        # Create a copy for styling (don't modify original)
        styled_df = scored.copy()
        # Ensure score columns are strings for consistent styling
        for col in score_columns_present:
            if col in styled_df.columns:
                # Convert to string, handling NaN properly
                styled_df[col] = styled_df[col].apply(
                    lambda x: "Missing" if pd.isna(x) else str(x).strip()
                )
        
        # Apply safe_for_display after preparing for styling
        styled_df = safe_for_display(styled_df)
        
        # Use map (newer pandas) or applymap (older pandas)
        try:
            styled = styled_df.style.map(highlight_scores, subset=score_columns_present)
        except (AttributeError, TypeError, ValueError):
            # Fallback for older pandas versions
            styled = styled_df.style.applymap(highlight_scores, subset=score_columns_present)
        
        st.dataframe(styled, use_container_width=True, height=600)
    except Exception as e:
        # If styling fails, show un-styled data with a warning
        st.warning(f"Could not apply color coding: {str(e)}. Showing data without styling.")
        st.dataframe(safe_for_display(scored), use_container_width=True, height=600)
else:
    st.dataframe(safe_for_display(scored), use_container_width=True, height=600)

# Downloads
st.markdown("---")
col1, col2 = st.columns(2)
with col1:
    st.download_button(
        "📥 Download Combined Raw Data (Excel)",
        data=to_excel_bytes(combined),
        file_name="combined_screeners.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
with col2:
    st.download_button(
        "📥 Download Scored Results (Excel)",
        data=to_excel_bytes(scored, score_columns=score_cols, raw_df=combined, criteria_df=criteria_df),
        file_name="scored_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# Create tabs for additional information
st.markdown("---")
tab1, tab2 = st.tabs(["📋 Ratio Criteria", "📊 Combined Data"])

with tab1:
    st.subheader("Ratio Criteria")
    st.caption("Scoring thresholds used for evaluation")
    st.dataframe(criteria_df[["factor", "description", "good", "avg", "bad"]], use_container_width=True)

with tab2:
    st.subheader("Combined Screener Data")
    st.caption(f"{len(combined)} tickers | {len(combined.columns)} columns after merge")
    st.dataframe(safe_for_display(combined.head(200)), use_container_width=True, height=500)

